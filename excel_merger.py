# -*- coding: utf-8 -*-
"""
Excel 多表合并工具
功能：多张 Excel 表按指定主键字段合并，支持字段级非空值合并
"""

import os
import sys
import threading
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import datetime


# =============================================================================
# 数据模型
# =============================================================================

class ExcelFile:
    """Excel文件对象"""
    def __init__(self, path):
        self.path = path
        self.name = os.path.basename(path)
        self.wb = None
        self.sheets = []
        self.selected_sheet = None
        self.headers = []
        self.data = []
        self._load_workbook()

    def _load_workbook(self):
        try:
            self.wb = openpyxl.load_workbook(self.path, data_only=True)
            self.sheets = self.wb.sheetnames
            if self.sheets:
                self.selected_sheet = self.sheets[0]
        except Exception as e:
            raise Exception(f"无法读取文件 {self.name}: {e}")

    def load_sheet(self, sheet_name):
        """加载指定sheet的数据"""
        ws = self.wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        headers = [str(h).strip() if h is not None else f"列{idx}" for idx, h in enumerate(rows[0])]
        data = rows[1:]
        self.selected_sheet = sheet_name
        self.headers = headers
        self.data = data
        return headers, data

    def close(self):
        if self.wb:
            self.wb.close()


class MergerLogic:
    """合并逻辑"""

    @staticmethod
    def merge_files(files, primary_key, progress_callback=None):
        """
        合并多个Excel文件
        files: ExcelFile列表
        primary_key: 主键字段名
        返回: (merged_headers, merged_data, stats_dict)
        """
        # 第1步：收集所有字段（按出现顺序，去重）
        all_field_order = []
        field_sources = {}  # field -> [(file_name, col_idx), ...]

        for f in files:
            for col_idx, header in enumerate(f.headers):
                if header not in field_sources:
                    field_sources[header] = []
                    all_field_order.append(header)
                field_sources[header].append((f.name, col_idx))

        # 第2步：建立主键索引
        # data_rows: {pk_value: {field: value}}
        pk_index = {}  # pk_value -> row_dict
        pk_row_count = {}  # pk_value -> 行数（不去重，每条保留）

        total_ops = sum(len(f.data) for f in files)
        completed_ops = [0]

        for f in files:
            pk_col_idx = None
            for idx, h in enumerate(f.headers):
                if h == primary_key:
                    pk_col_idx = idx
                    break

            if pk_col_idx is None:
                raise Exception(f"文件 {f.name} 中找不到主键字段「{primary_key}」")

            for row in f.data:
                completed_ops[0] += 1
                progress_callback and progress_callback(completed_ops[0], total_ops)

                pk_val = row[pk_col_idx]
                pk_str = str(pk_val) if pk_val is not None else "__NULL__"

                if pk_str not in pk_index:
                    pk_index[pk_str] = {"_pk_val": pk_val}
                    pk_row_count[pk_str] = 0
                else:
                    old_pk = pk_index[pk_str].get("_pk_val")
                    if old_pk is None or (pk_val is not None and str(pk_val).strip() != ""):
                        pk_index[pk_str]["_pk_val"] = pk_val

                pk_row_count[pk_str] += 1

                # 写入每个字段（取非空值）
                for col_idx, header in enumerate(f.headers):
                    cell_val = row[col_idx]
                    # 非空判断
                    if cell_val is not None and str(cell_val).strip() != "":
                        pk_index[pk_str][header] = cell_val

        # 第3步：生成合并后数据
        merged_data = []
        for pk_val in pk_index:
            row_dict = {k: v for k, v in pk_index[pk_val].items() if k != "_pk_val"}
            row_dict[primary_key] = pk_index[pk_val]["_pk_val"]
            merged_data.append(row_dict)

        # 统计信息
        stats = {
            "total_files": len(files),
            "total_rows_before": sum(len(f.data) for f in files),
            "total_rows_after": len(merged_data),
            "pk_duplicates": sum(1 for v in pk_row_count.values() if v > 1),
            "fields_count": len(all_field_order),
        }

        return all_field_order, merged_data, stats


# =============================================================================
# GUI 应用
# =============================================================================

class ExcelMergerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 多表合并工具 v1.0")
        self.root.geometry("900x700")
        self.root.configure(bg="#f0f0f0")

        self.files = []          # ExcelFile对象列表
        self.all_fields = []     # 所有文件的所有字段
        self.primary_key = None
        self.merged_headers = []
        self.merged_data = []

        self._build_ui()

    def _build_ui(self):
        """构建界面"""
        # ----- 顶部标题 -----
        header_frame = Frame(self.root, bg="#2c3e50", height=50)
        header_frame.pack(fill=X)
        header_frame.pack_propagate(False)
        Label(header_frame, text="Excel 多表合并工具", font=("微软雅黑", 16, "bold"),
              fg="white", bg="#2c3e50").pack(pady=12)

        # ----- 主内容区 -----
        main_frame = Frame(self.root, bg="#f0f0f0")
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        # --- 左侧：文件列表 ---
        left_frame = Frame(main_frame, bg="white")
        left_frame.pack(side=LEFT, fill=Y, padx=(0, 10))
        Label(left_frame, text="已添加文件", font=("微软雅黑", 11, "bold"),
              bg="white").pack(pady=(10, 5))

        list_frame = Frame(left_frame, bg="white")
        list_frame.pack(fill=Y, expand=True, padx=10)

        self.file_listbox = Listbox(list_frame, width=35, height=20,
                                    font=("微软雅黑", 10), selectmode=EXTENDED)
        self.file_listbox.pack(side=LEFT, fill=Y)
        scroll_file = Scrollbar(list_frame, command=self.file_listbox.yview)
        scroll_file.pack(side=RIGHT, fill=Y)
        self.file_listbox.configure(yscrollcommand=scroll_file.set)

        btn_frame = Frame(left_frame, bg="white")
        btn_frame.pack(pady=10, padx=10)
        Button(btn_frame, text="添加文件", command=self.add_files,
               bg="#27ae60", fg="white", font=("微软雅黑", 10), width=10).pack(side=LEFT, padx=3)
        Button(btn_frame, text="删除选中", command=self.remove_selected,
               bg="#e74c3c", fg="white", font=("微软雅黑", 10), width=10).pack(side=LEFT, padx=3)

        # --- 中间：Sheet选择 & 字段显示 ---
        mid_frame = Frame(main_frame, bg="#f0f0f0")

        # Sheet选择区
        sheet_label = Label(mid_frame, text="选择每张表的Sheet", font=("微软雅黑", 11, "bold"),
                            bg="#f0f0f0")
        sheet_label.pack(pady=(0, 5))

        self.sheet_container = Frame(mid_frame, bg="#f0f0f0")
        self.sheet_container.pack(fill=X, pady=(0, 10))

        # 字段预览区
        field_label = Label(mid_frame, text="所有字段列表", font=("微软雅黑", 11, "bold"),
                            bg="#f0f0f0")
        field_label.pack(pady=(0, 5))

        field_list_frame = Frame(mid_frame, bg="white")
        field_list_frame.pack(fill=X)

        self.field_listbox = Listbox(field_list_frame, width=30, height=8,
                                     font=("微软雅黑", 9), selectmode=BROWSE)
        self.field_listbox.pack(side=LEFT, fill=X, expand=True)
        scroll_field = Scrollbar(field_list_frame, command=self.field_listbox.yview)
        scroll_field.pack(side=RIGHT, fill=Y)
        self.field_listbox.configure(yscrollcommand=scroll_field.set)

        # 主键选择
        pk_frame = Frame(mid_frame, bg="#f0f0f0")
        pk_frame.pack(fill=X, pady=(15, 5))
        Label(pk_frame, text="选择主键字段：", font=("微软雅黑", 11),
              bg="#f0f0f0").pack(side=LEFT)
        self.pk_var = StringVar()
        self.pk_combo = ttk.Combobox(pk_frame, textvariable=self.pk_var,
                                      font=("微软雅黑", 10), width=20, state="readonly")
        self.pk_combo.pack(side=LEFT, padx=5)
        self.pk_combo.bind("<<ComboboxSelected>>", self.on_pk_selected)

        mid_frame.pack(side=LEFT, fill=Y, padx=(0, 10))

        # --- 右侧：预览 & 操作 ---
        right_frame = Frame(main_frame, bg="white")

        # 预览区标题
        Label(right_frame, text="合并预览", font=("微软雅黑", 11, "bold"),
              bg="white").pack(pady=(10, 5))

        # 统计信息
        self.stats_label = Label(right_frame, text="请先添加文件并选择主键",
                                  font=("微软雅黑", 9), bg="white", fg="#7f8c8d",
                                  anchor=W, justify=LEFT)
        self.stats_label.pack(fill=X, padx=10)

        # 预览表格（Treeview）
        preview_frame = Frame(right_frame)
        preview_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)

        self.preview_tree = ttk.Treeview(preview_frame, show="headings", height=15)
        preview_tree_scroll_y = Scrollbar(preview_frame, command=self.preview_tree.yview)
        preview_tree_scroll_x = Scrollbar(preview_frame, orient=HORIZONTAL,
                                           command=self.preview_tree.xview)
        self.preview_tree.configure(yscrollcommand=preview_tree_scroll_y.set,
                                     xscrollcommand=preview_tree_scroll_x.set)
        self.preview_tree.grid(row=0, column=0, sticky=N+S+E+W)
        preview_tree_scroll_y.grid(row=0, column=1, sticky=N+S)
        preview_tree_scroll_x.grid(row=1, column=0, sticky=E+W)
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)

        # 操作按钮
        op_frame = Frame(right_frame, bg="white")
        op_frame.pack(pady=10, padx=10)

        self.merge_btn = Button(op_frame, text="执行合并", command=self.do_merge,
                                 bg="#2980b9", fg="white", font=("微软雅黑", 11), width=12,
                                 state=DISABLED)
        self.merge_btn.pack(side=LEFT, padx=5)

        Button(op_frame, text="导出Excel", command=self.export_excel,
               bg="#8e44ad", fg="white", font=("微软雅黑", 11), width=12,
               state=DISABLED).pack(side=LEFT, padx=5)

        self.export_btn_state = DISABLED

        right_frame.pack(side=LEFT, fill=BOTH, expand=True)

        # ----- 底部状态栏 -----
        self.status_var = StringVar()
        self.status_var.set("就绪")
        status_bar = Label(self.root, textvariable=self.status_var,
                           font=("微软雅黑", 9), bd=1, relief=SUNKEN, anchor=W)
        status_bar.pack(side=BOTTOM, fill=X)

        # 进度条
        self.progress_frame = Frame(self.root, bg="#f0f0f0", height=30)
        self.progress_frame.pack(fill=X, padx=10, pady=(0, 5))
        self.progress_frame.pack_propagate(False)
        self.progress_label = Label(self.progress_frame, text="", font=("微软雅黑", 9),
                                     bg="#f0f0f0", anchor=W)
        self.progress_label.pack(side=LEFT)
        self.progress_bar = ttk.Progressbar(self.progress_frame, length=300, mode="determinate")
        self.progress_bar.pack(side=RIGHT)

    # -------------------------------------------------------------------------
    # 文件操作
    # -------------------------------------------------------------------------

    def add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择Excel文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not paths:
            return

        for path in paths:
            # 查重
            if any(f.path == path for f in self.files):
                messagebox.showwarning("重复文件", f"已添加：{os.path.basename(path)}")
                continue
            try:
                ef = ExcelFile(path)
                self.files.append(ef)
                self.file_listbox.insert(END, ef.name)
            except Exception as e:
                messagebox.showerror("错误", str(e))

        self.status_var.set(f"已添加 {len(self.files)} 个文件")
        self._refresh_sheet_ui()
        self._load_all_fields()

    def remove_selected(self):
        sel = self.file_listbox.curselection()
        if not sel:
            return
        # 从后往前删
        for idx in reversed(sel):
            self.files[idx].close()
            del self.files[idx]
            self.file_listbox.delete(idx)

        self._refresh_sheet_ui()
        self._load_all_fields()

    # -------------------------------------------------------------------------
    # Sheet 选择 UI（动态构建）
    # -------------------------------------------------------------------------

    def _refresh_sheet_ui(self):
        """刷新Sheet选择区域"""
        for w in self.sheet_container.winfo_children():
            w.destroy()

        if not self.files:
            return

        for idx, f in enumerate(self.files):
            row = Frame(self.sheet_container, bg="white")
            row.pack(fill=X, pady=2)

            name_label = Label(row, text=f.name[:20], font=("微软雅黑", 9),
                               bg="white", width=20, anchor=W)
            name_label.pack(side=LEFT, padx=(5, 3))

            sheet_var = StringVar(value=f.selected_sheet)
            combo = ttk.Combobox(row, textvariable=sheet_var,
                                  values=f.sheets, font=("微软雅黑", 9),
                                  width=15, state="readonly")
            combo.pack(side=LEFT, padx=3)

            def on_sheet_change(e, file_idx=idx, combo_ref=combo):
                new_sheet = combo_ref.get()
                self.files[file_idx].load_sheet(new_sheet)
                self._load_all_fields()

            combo.bind("<<ComboboxSelected>>", on_sheet_change)
            # 自动触发加载第一张sheet
            if f.selected_sheet and not f.headers:
                f.load_sheet(f.selected_sheet)

    # -------------------------------------------------------------------------
    # 字段 & 主键
    # -------------------------------------------------------------------------

    def _load_all_fields(self):
        """收集所有文件的所有字段"""
        self.all_fields = []
        self.field_listbox.delete(0, END)

        for f in self.files:
            for h in f.headers:
                if h not in self.all_fields:
                    self.all_fields.append(h)
                    self.field_listbox.insert(END, h)

        # 更新主键下拉框
        self.pk_combo["values"] = self.all_fields
        if self.all_fields:
            self.pk_var.set("")
        else:
            self.pk_var.set("")

    def on_pk_selected(self, event=None):
        self.primary_key = self.pk_var.get()
        if self.primary_key and self.files:
            self.merge_btn.config(state=NORMAL)

    # -------------------------------------------------------------------------
    # 合并 & 预览
    # -------------------------------------------------------------------------

    def _update_preview(self, headers, data, max_rows=50):
        """更新预览表格"""
        self.preview_tree.delete(*self.preview_tree.get_children())
        for col in self.preview_tree["columns"]:
            self.preview_tree.heading(col, text="")
            self.preview_tree.column(col, width=0)

        if not headers:
            return

        # 设置列
        self.preview_tree["columns"] = headers
        col_width = min(120, max(60, 800 // len(headers)))
        for h in headers:
            self.preview_tree.heading(h, text=h)
            self.preview_tree.column(h, width=col_width, anchor=W)

        # 插入数据
        for row_data in data[:max_rows]:
            values = [row_data.get(h, "") for h in headers]
            self.preview_tree.insert("", END, values=values)

    def do_merge(self):
        if not self.primary_key or not self.files:
            messagebox.showwarning("条件不足", "请选择主键字段")
            return

        self.status_var.set("正在合并...")
        self.merge_btn.config(state=DISABLED, text="合并中...")

        def merge_thread():
            try:
                logic = MergerLogic()
                headers, data, stats = logic.merge_files(
                    self.files,
                    self.primary_key,
                    progress_callback=self._on_merge_progress
                )

                self.merged_headers = headers
                self.merged_data = data

                # 更新UI（主线程）
                def update_ui():
                    # 预览
                    self._update_preview(headers, data)
                    # 统计
                    stats_text = (
                        f"文件数：{stats['total_files']}　"
                        f"合并前总行数：{stats['total_rows_before']}　"
                        f"合并后行数：{stats['total_rows_after']}　"
                        f"主键重复行：{stats['pk_duplicates']}　"
                        f"字段数：{stats['fields_count']}"
                    )
                    self.stats_label.config(text=stats_text)
                    # 启用导出
                    for w in self.root.winfo_children():
                        pass
                    self._enable_export()

                    self.status_var.set("合并完成")
                    self.merge_btn.config(state=NORMAL, text="执行合并")
                    messagebox.showinfo("完成", f"合并完成！\n合并后 {stats['total_rows_after']} 行，{stats['fields_count']} 个字段")

                self.root.after(0, update_ui)

            except Exception as e:
                def on_error():
                    self.status_var.set("合并失败")
                    self.merge_btn.config(state=NORMAL, text="执行合并")
                    messagebox.showerror("错误", str(e))
                self.root.after(0, on_error)

        threading.Thread(target=merge_thread, daemon=True).start()

    def _on_merge_progress(self, current, total):
        def update_progress():
            pct = int(current / total * 100)
            self.progress_bar["value"] = pct
            self.progress_label.config(text=f"正在合并... {current}/{total} ({pct}%)")
        self.root.after(0, update_progress)

    def _enable_export(self):
        self.export_btn_state = NORMAL
        # 重新查找export按钮并启用
        for w in self.root.winfo_children():
            if isinstance(w, Frame):
                for child in w.winfo_children():
                    if isinstance(child, Frame):
                        for btn in child.winfo_children():
                            if isinstance(btn, Button) and btn.cget("text") == "导出Excel":
                                btn.config(state=NORMAL)

    # -------------------------------------------------------------------------
    # 导出
    # -------------------------------------------------------------------------

    def export_excel(self):
        if not self.merged_data:
            messagebox.showwarning("无数据", "请先执行合并")
            return

        path = filedialog.asksaveasfilename(
            title="导出合并结果",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")]
        )
        if not path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "合并结果"

            # 样式定义
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            data_align = Alignment(horizontal="left", vertical="center")

            # 写入表头
            for col_idx, h in enumerate(self.merged_headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            # 写入数据
            for row_idx, row_dict in enumerate(self.merged_data, 2):
                for col_idx, field in enumerate(self.merged_headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_dict.get(field, ""))
                    cell.alignment = data_align
                    cell.border = border

            # 设置列宽
            for col_idx, h in enumerate(self.merged_headers, 1):
                col_letter = get_column_letter(col_idx)
                max_len = max(len(str(h)), 30)
                ws.column_dimensions[col_letter].width = min(max_len, 50)

            # 冻结首行
            ws.freeze_panes = "A2"

            wb.save(path)
            self.status_var.set(f"导出成功：{os.path.basename(path)}")
            messagebox.showinfo("成功", f"已导出：\n{path}")

        except Exception as e:
            messagebox.showerror("导出失败", str(e))


# =============================================================================
# 入口
# =============================================================================

def main():
    root = Tk()
    app = ExcelMergerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()